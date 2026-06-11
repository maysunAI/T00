"""
excel_to_md.py — Excel 設計書 → Markdown 変換ツール
使い方: python excel_to_md.py <input.xlsx> [output.md]
"""
import sys
import os

def check_deps():
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl")
        import openpyxl
    return openpyxl

def cell_value(cell):
    if cell.value is None:
        return ""
    return str(cell.value).strip().replace("\n", "<br>")

def sheet_to_markdown(ws):
    rows = list(ws.iter_rows())
    if not rows:
        return ""

    lines = []
    lines.append(f"## シート：{ws.title}\n")

    # Get column widths for alignment
    max_col = max(len(row) for row in rows)

    # Build table
    table_rows = []
    for row in rows:
        cells = [cell_value(c) for c in row]
        # Pad to max_col
        while len(cells) < max_col:
            cells.append("")
        table_rows.append(cells)

    if not table_rows:
        return ""

    # Header row
    header = table_rows[0]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join(["---"] * max_col) + "|")

    # Data rows
    for row in table_rows[1:]:
        # Skip completely empty rows
        if all(c == "" for c in row):
            continue
        lines.append("| " + " | ".join(row) + " |")

    lines.append("")
    return "\n".join(lines)

def excel_to_markdown(xlsx_path, output_path=None):
    openpyxl = check_deps()

    if not os.path.exists(xlsx_path):
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    md_parts = []
    md_parts.append(f"# 設計書：{os.path.basename(xlsx_path)}\n")
    md_parts.append(f"> 変換元：{xlsx_path}  \n> シート数：{len(wb.sheetnames)}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_section = sheet_to_markdown(ws)
        if md_section:
            md_parts.append(md_section)
            md_parts.append("---\n")

    md_content = "\n".join(md_parts)

    if output_path is None:
        base = os.path.splitext(xlsx_path)[0]
        output_path = base + ".md"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"Converted: {xlsx_path}")
    print(f"Output:    {output_path}")
    print(f"Sheets:    {len(wb.sheetnames)}")
    print(f"Lines:     {len(md_content.splitlines())}")
    return output_path

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_md.py <input.xlsx> [output.md]")
        print("")
        print("Example:")
        print("  python excel_to_md.py design_spec.xlsx")
        print("  python excel_to_md.py design_spec.xlsx output/design.md")
        sys.exit(0)

    xlsx_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    excel_to_markdown(xlsx_file, out_file)
